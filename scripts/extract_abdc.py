#!/usr/bin/env python3
"""Extract A* and A journals from the ABDC Journal Quality List xlsx."""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl first: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"


def find_xlsx():
    files = list(DATA_DIR.glob("ABDC*.xlsx"))
    if not files:
        print(f"No ABDC xlsx found in {DATA_DIR}")
        print("Download from https://abdc.edu.au and place in data/")
        sys.exit(1)
    if len(files) > 1:
        print(f"Multiple ABDC xlsx files found: {[f.name for f in files]}")
        print("Keep only one.")
        sys.exit(1)
    return files[0]


def extract(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    title_col = None
    rating_col = None
    for i, h in enumerate(header):
        if "title" in h and title_col is None:
            title_col = i
        if "rating" in h or "rank" in h:
            rating_col = i

    if title_col is None or rating_col is None:
        print(f"Could not find title/rating columns in header: {header}")
        sys.exit(1)

    print(f"Using column {title_col} ('{header[title_col]}') for title")
    print(f"Using column {rating_col} ('{header[rating_col]}') for rating")

    journals = []
    for row in rows[1:]:
        if not row[title_col] or not row[rating_col]:
            continue
        name = str(row[title_col]).strip()
        rating = str(row[rating_col]).strip().upper()
        if rating in ("A*", "A"):
            journals.append({"name": name, "rating": rating})

    journals.sort(key=lambda j: j["name"].lower())
    return journals


def main():
    xlsx_path = find_xlsx()
    print(f"Reading: {xlsx_path.name}")

    journals = extract(xlsx_path)
    a_star = sum(1 for j in journals if j["rating"] == "A*")
    a_only = sum(1 for j in journals if j["rating"] == "A")
    print(f"  A*: {a_star}")
    print(f"  A:  {a_only}")
    print(f"  Total: {len(journals)}")

    output_path = DATA_DIR / "abdc.json"
    with open(output_path, "w") as f:
        json.dump(journals, f, indent=2, ensure_ascii=False)
    print(f"Written to {output_path}")


if __name__ == "__main__":
    main()
